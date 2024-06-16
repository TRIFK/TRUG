from django.shortcuts import redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User, Group
from .forms import RegistrationForm
from django.contrib.auth.forms import AuthenticationForm
from .models import Product, ProductLocation, Supply, SupplyProduct, Shipment, ShipmentProduct, Order, OrderProduct, ProductType
from django.http import JsonResponse
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import date
from django.shortcuts import render
from django.views.decorators.http import require_POST, require_GET
from django.core.cache import cache
from django.contrib import messages
from .forms import OrderForm
import json
from django.views.decorators.csrf import csrf_exempt




def welcome(request):
    messages.info(request, 'Ваша сессия истекла. Пожалуйста, войдите в систему снова.')
    redirect('/welcome/')
    return render(request, 'welcome.html')

def auth(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                return render(request, 'auth.html', {'form': form, 'error': 'Неверные учетные данные'})
        else:
            return render(request, 'auth.html', {'form': form, 'error': 'Неверные учетные данные'})
    else:
        form = AuthenticationForm()
    return render(request, 'auth.html', {'form': form})

def in_group(group_name):
    def check(user):
        return user.groups.filter(name=group_name).exists() or user.is_superuser
    return check

def logout_view(request):
    logout(request)
    return redirect('welcome')


def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=False)
                user.save()
                group = form.cleaned_data.get('group')
                user.groups.add(group)
                user = authenticate(username=form.cleaned_data['username'], password=form.cleaned_data['password1'])
                if user is not None:
                    login(request, user)
                    return redirect('home')
                else:
                    return render(request, 'register.html', {'form': form, 'error': 'Ошибка аутентификации'})
            except Exception as e:
                return render(request, 'register.html', {'form': form, 'error': f'Ошибка при сохранении пользователя: {str(e)}'})
        else:
            groups = Group.objects.all()

            # Вывод ошибок формы для отладки
            print(form.errors)
            return render(request, 'register.html', {'form': form, 'error': 'Форма не валидна', 'groups': groups})
    else:
        form = RegistrationForm()

    groups = Group.objects.all()
    context = {'form': form, 'groups': groups}
    return render(request, 'register.html', context)

def home(request):
    orders = Order.objects.all()
    orders_products = OrderProduct.objects.all()

    # Попытка получить ближайшую дату отгрузки из кеша Redis
    nearest_shipment_date = cache.get('nearest_shipment_date')

    if not nearest_shipment_date:
        # Если дата не найдена в кеше, выполним запрос к базе данных
        try:
            nearest_shipment = Shipment.objects.filter(date_shipped__gte=date.today()).earliest('date_shipped')
            nearest_shipment_date = nearest_shipment.date_shipped
            # Сохраняем результат в кеш Redis на 15 минут
            cache.set('nearest_shipment_date', nearest_shipment_date, timeout=900)  # 900 секунд = 15 минут
        except Shipment.DoesNotExist:
            nearest_shipment_date = None

    context = {
        'orders': orders,
        'nearest_shipment_date': nearest_shipment_date,
    }
    return render(request, 'home.html', context)


def products(request):
    products = Product.objects.all()
    context = {
        'products': products
    }
    return render(request, 'products.html', context)

def placement(request):
    product_locations = ProductLocation.objects.all()
    context = {
        'product_locations': product_locations
    }
    return render(request, 'placement.html', context)

def supplies(request):
    supplies = Supply.objects.all()
    supplies_products = SupplyProduct.objects.all()
    context = {
        'supplies': supplies,
        'supplies_products': supplies_products
    }
    return render(request, 'supplies.html', context)

def shipment(request):
    shipments = Shipment.objects.all()
    shipment_products = ShipmentProduct.objects.all()
    context = {
        'shipments': shipments,
        'shipment_products': shipment_products
    }
    return render(request, 'shipment.html', context)


def generate_invoice(request):
    # Fetch all orders
    orders = Order.objects.all()

    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'order'

    # Define headers for the Excel sheet
    headers = ['Клиент', 'Продукция', 'Сумма заказа', 'Дата заказа']

    # Write headers into the first row of the sheet
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write order data into subsequent rows
    for row_num, order in enumerate(orders, 2):  # Start from row 2 for data rows
        products_list = ', '.join([str(op.product) for op in order.products.all()])
        data = [order.customer, products_list, order.summary, order.date_ordered]

        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Save the Excel file to a BytesIO object
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Set up the HTTP response
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openpyxl.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'

    return response

@require_GET
def orders(request):
    orders = Order.objects.all()
    orders_products = OrderProduct.objects.all()
    products = Product.objects.all()
    product_types = ProductType.objects.all()

    # Преобразуем данные о продуктах в JSON-совместимый формат
    products_data = list(products.values('id', 'name'))

    context = {
        'orders': orders,
        'orders_products': orders_products,
        'products': products_data,  # Передаем только данные о продуктах
        'product_types': product_types
    }
    return render(request, 'orders.html', context)


@csrf_exempt
def create_order(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            data['completed'] = False

            products = data.get('selected_products', [])

            form_data = {
                'customer': data.get('customer', ''),
                'summary': data.get('summary', 0),
                'date_ordered': data.get('date_ordered', ''),
            }

            # Сначала создаем заказ (Order)
            form = OrderForm(form_data)
            if form.is_valid():
                order_instance = form.save(commit=False)
                order_instance.completed = False  # Устанавливаем completed на False
                order_instance.save()

                ids_ord = []
                # Теперь добавляем продукты в заказ (OrderProduct)
                for product_data in products:
                    product_id = product_data['ID']
                    quantity = product_data['quantity']

                    order_product = OrderProduct.objects.create(product_id=product_id, quantity=quantity)
                    ids_ord.append(order_product.id)

                order_instance.products.set(ids_ord)
                return JsonResponse({'success': True})
            else:
                errors = form.errors.as_json()
                return JsonResponse({'success': False, 'error': errors}, status=400)

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)




@csrf_exempt
@require_POST
def delete_orders(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            order_ids = data.get('order_id', [])
            if order_ids and isinstance(order_ids, list):
                # Удаление заказов из базы данных
                deleted_count, _ = Order.objects.filter(id__in=order_ids).delete()
                if deleted_count > 0:
                    return JsonResponse({'success': True, 'message': 'Заказы успешно удалены'}, status=200)
                else:
                    return JsonResponse({'success': False, 'message': 'Заказы не найдены'}, status=404)
            else:
                return JsonResponse({'success': False, 'message': 'Некорректные данные'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Некорректный формат JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Неправильный тип запроса'}, status=400)


@csrf_exempt
@require_POST
def edit_order(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            order_id = data.get('order_id')
            customer = data.get('customer')
            summary = data.get('summary')
            date_ordered = data.get('date_ordered')
            products = data.get('products', [])

            if not order_id or not customer or not summary or not date_ordered or not products:
                return JsonResponse({'success': False, 'message': 'Некорректные данные'}, status=400)

            try:
                order = Order.objects.get(id=order_id)
            except Order.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Заказ не найден'}, status=404)

            order.customer = customer
            order.summary = summary
            order.date_ordered = date_ordered
            order.save()

            order.products.clear()
            for product_data in products:
                product_id = product_data['id']
                quantity = product_data['quantity']
                order_product = OrderProduct.objects.create(product_id=product_id, quantity=quantity)
                order.products.add(order_product)

            order.save()

            return JsonResponse({'success': True, 'order': {
                'id': order.id,
                'customer': order.customer,
                'summary': order.summary,
                'date_ordered': order.date_ordered,
                'products': [{'name': op.product.name, 'quantity': op.quantity} for op in order.products.all()]
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Некорректный формат JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Неправильный тип запроса'}, status=400)



